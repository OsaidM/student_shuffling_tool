import config as SCon
import argparser
import pandas
import random
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.reader.excel import load_workbook

#################################################################
## index file where all the shuffling process will be handeled ##
#################################################################

#########################
## important instances ##
######################### 
wb = Workbook()


class Students:
    '''
    Student class responsible for Shuffling processes 
    '''
    def is_none(arg):
        '''
        Takse any variable and returns True or False if contains any value but None
        '''
        if arg == None:
            return True
        return False
    
    def initiate_excel_sheet(name):
        '''
        Takes a string name and create excel file based on that name with one empty sheet 
        '''
        wb.save(f'{name}.xlsx')
        return print("Initializing Sections ...")

    def merge_two_lists_to_tuples_list(current_dict, colname1, colname2):
        '''
        takes a python dictinary and columns names to read from \
        return one list of tuples merged out of two lists
        '''
        for key in current_dict[f"{colname1}"]:
            SCon.extracted_column_1.append(current_dict[f"{colname1}"][key])
        for key in current_dict[f"{colname2}"]:
            SCon.extracted_column_2.append(current_dict[f"{colname2}"][key])
        for i in range(len(current_dict["Academy ID"])):
            SCon.students_list.append(tuple((SCon.extracted_column_1[i], SCon.extracted_column_2[i])))
        return SCon.students_list
    
    def remove_duplicates(section, current_list):
        '''
        takes a python dictinary and returns a list of shuffled students
        '''
        for x in section:
            current_list.remove(x)

    def initialize_section(number_of_sections):
        '''
        intializes the sections inside the shuffled_sections_dict
        '''
        for i in range(number_of_sections):
            SCon.shuffled_sections_dict[f'section{i+1}'] = []
        return SCon.shuffled_sections_dict

    def get_maximum_array_length(shffled_sections):
        '''
        takes a shffled_sections dictionary and checks for the length is less than maximum
        and returns the maximum_array_length
        '''
        maximum_array_length = 0
        for key in shffled_sections:
            if len(shffled_sections[key]) > maximum_array_length:
                maximum_array_length = len(shffled_sections[key])
        return maximum_array_length

    def balance_section_arrays(shffled_sections):
        '''
        takes a shffled_sections dictionary and adding empty cells inside each array
        if the length is less than the maximum
        '''
        maximum_array_length = Students.get_maximum_array_length(shffled_sections)
        for key in shffled_sections:
            if len(shffled_sections[key]) < maximum_array_length:
                shffled_sections[key].append(tuple(('','')))

    def read_from_excel(xcel_file, column_name):
        '''
        takes an excel file path as string and returns a students list of tuples 
        '''
        try:
            captured_excel = pandas.read_excel(f'{xcel_file}.xlsx', engine='openpyxl')
            df = pandas.DataFrame(captured_excel, columns=[f'{column_name}'])
            current_dict = df[f'{column_name}'].to_dict()
            for key in current_dict:
                SCon.students_list.append(tuple((key,current_dict[key])))
            return SCon.students_list
        except FileNotFoundError:
            print("\n \n File does not exist!")
            
        except:
            print("\n \n Excel file is not correct")
            
    
    def read_from_excel2(xcel_file, column_names):
        '''
        takes an excel file path as string and returns a students list of tuples 
        '''
        print(column_names)
        try:
            captured_excel = pandas.read_excel(f'{xcel_file}.xlsx', engine='openpyxl')
            df = pandas.DataFrame(captured_excel, columns=column_names)
            current_dict = df.to_dict()
            SCon.students_list = Students.merge_two_lists_to_tuples_list(current_dict, column_names[0], column_names[1])
            return SCon.students_list
        except FileNotFoundError:
            print("\n \n File does not Exist!")
            
        except:
            print("\n \n Excel file is not Correct")
            

    def students_per_section_shuffle(file_path, sections_number, column_name, column_names):
        '''
        takes a python dictinary and returns a list of shuffled students
        '''
        if len(column_names)>0:
            current_list = Students.read_from_excel2(file_path, column_names)
        else:
            current_list = Students.read_from_excel(file_path, column_name)
        if Students.is_none(current_list):
            exit()    
        students_number = len(current_list)
        while students_number > 0 and sections_number > 0:
            section = random.sample(current_list, int(students_number/sections_number))
            Students.remove_duplicates(section, current_list)
            SCon.shuffled_sections_dict[f'section{sections_number}'] = section
            Students.balance_section_arrays(SCon.shuffled_sections_dict)
            students_number -= int(students_number/sections_number)
            sections_number -= 1
        return SCon.shuffled_sections_dict
    
    def apply_style_for_each_sheet(excel_name):
        '''
        Takes Excel name as string and loades the excel file then apply the style for each sheet
        '''
        workbook = load_workbook(f'{excel_name}.xlsx')
        for i in range(len(workbook.sheetnames)):
            first_cell = workbook.worksheets[i].cell(1,1)
            second_cell = workbook.worksheets[i].cell(1,2)
            first_cell.font = Font(name='Arial', size=22, bold=True)
            second_cell.font = Font(name='Arial', size=22, bold=True)
            first_cell.alignment = Alignment(horizontal="center", vertical="center", shrink_to_fit=True)
            workbook.worksheets[i].column_dimensions["B"].width = 20
        workbook.save(f'{excel_name}.xlsx')
        return print("Applyed Styles.")

    def generate_excel_sheet_option1(shuffled_list, excel_name, generated_columns_names, styling):
        '''
        takes a shuffled list of stacks for each section and generate excel sheet based on that and return current_stack  
        '''
        Students.initiate_excel_sheet(excel_name)
        for i in range(len(shuffled_list)):
            for key in shuffled_list[i][f'stack{i}']:
                current_section = pandas.DataFrame.from_dict(data=shuffled_list[i][f'stack{i}'][key])
                if len(generated_columns_names) > 0:
                    current_section.columns = generated_columns_names
                with pandas.ExcelWriter(f'{excel_name}.xlsx', engine="openpyxl", mode='a') as writer:
                    current_section.to_excel(writer, sheet_name=f'{shuffled_list[i]["stack_name"]}, {key}', index=False)
        if styling:
            Students.apply_style_for_each_sheet(excel_name)
        return print('\n \n Generated Excel Sheet Successfully!')

    def generate_excel_sheet_option2(shuffled_list, excel_name, generated_columns_names, styling):
        '''
        takes a shuffled list of stacks for each section and generate excel sheet based on that and return current_stack  
        '''
        Students.initiate_excel_sheet(excel_name)
        for i in range(len(shuffled_list)):
            current_stack = pandas.DataFrame.from_dict(shuffled_list[i][f'stack{i}'])
            with pandas.ExcelWriter(f'{excel_name}.xlsx', engine="openpyxl", mode='a') as writer:
                current_stack.to_excel(writer, sheet_name=shuffled_list[i]['stack_name'], index=False)
        if styling:
            Students.apply_style_for_each_sheet(excel_name)
        return print('\n \n Generated Excel Sheet Successfully!')

    def check_errors(suffletype, generated_columns_names):
        '''
        Validation function Reads Paramters and check for expected errors and exit.
        this function expected to hold all possible combos error messages.
        '''
        if len(generated_columns_names) > 0 and suffletype == 2:
            print("\n \n Error: -wclsl command works only with shuffling type '1'.")
            exit()

    def students_shuffle(file_path, sections_number, stacks_list, excel_name, suffletype, column_name, read_column_names, generated_columns_names, styling):
        '''
        takes a file_path as string and sections_number as integer and returns generate_excel_sheet function
        or print error and exit.
        '''
        Students.check_errors(suffletype, generated_columns_names)
        Students.initialize_section(int(sections_number))
        print('Initialized Sections Successfully')
        print("Starts Reading and Shuffling...")
        for i in range(len(stacks_list)):
            this_stack_shuffle = Students.students_per_section_shuffle(file_path, int(sections_number), column_name, read_column_names)
            SCon.stacks_shuffled.append({'stack_name':stacks_list[i],f'stack{i}':this_stack_shuffle.copy()})
        print("Finished Shuffling.")
        if suffletype == 1:
            return Students.generate_excel_sheet_option1(SCon.stacks_shuffled, excel_name, generated_columns_names, styling)
        if suffletype == 2:
            return Students.generate_excel_sheet_option2(SCon.stacks_shuffled, excel_name, generated_columns_names, styling)
        else:
            print("\n \n Error: Please Enter a valid type option!")
            exit()


args = argparser.parser.parse_args()
if args.path:
    Students.students_shuffle(args.path, args.sections, args.stacks, args.generatedexcelname, args.shuffletype, args.rcolname, args.rcolslist, args.wcolsl, args.withstyle)
else:
    Students.students_shuffle(args.file, args.sections, args.stacks, args.generatedexcelname, args.shuffletype, args.rcolname, args.rcolslist, args.wcolsl, args.withstyle)

